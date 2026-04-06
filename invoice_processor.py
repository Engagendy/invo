import argparse
import csv
import hashlib
import importlib.util
import os
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, List, Optional, Sequence, Tuple

import cv2
import fitz
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageDraw, ImageFont
import pytesseract

from app_paths import bundled_models_dir, runtime_root, user_data_root


BoundingBox = Tuple[int, int, int, int]


@dataclass
class DocumentFields:
    doc_type: str = "Unknown"
    date: str = "Unknown"
    number: str = "Unknown"
    company_name: str = "Unknown"
    amount: str = "Unknown"


@dataclass
class ProcessedRecord:
    source_file: str
    source_path: str
    source_hash: str
    source_type: str
    source_origin: str
    source_timestamp: str
    output_file: str
    doc_type: str
    date: str
    number: str
    company_name: str
    amount: str
    currency: str
    transaction_direction: str
    project_name: str
    confidence_score: int
    confidence_label: str
    raw_text: str
    match_status: str = "unreviewed"
    match_score: int = 0
    matched_record_source_file: str = ""
    matched_record_output_file: str = ""
    matched_record_source_type: str = ""
    matched_record_source_timestamp: str = ""
    matched_record_date: str = ""
    matched_record_number: str = ""
    matched_record_company_name: str = ""
    matched_record_amount: str = ""
    matched_record_transaction_direction: str = ""
    match_basis: str = ""


DEFAULT_NAMING_PATTERN = (
    "{doc_type}_{date}_{number}_{company_name}_{amount_aed}_{project_name}"
)


def resolve_base_dir() -> Path:
    return runtime_root()


def resolve_model_dirs() -> List[Path]:
    return [
        user_data_root() / "models" / "official_models",
        bundled_models_dir() / "official_models",
        Path.home() / ".paddlex" / "official_models",
    ]


def trocr_cache_root() -> Path:
    return user_data_root() / "models" / "huggingface"


def trocr_models_root() -> Path:
    return user_data_root() / "models" / "trocr"


def trocr_model_dir(model_name: str) -> Path:
    return trocr_models_root() / model_name.replace("/", "--")


def bundled_trocr_model_dir(model_name: str) -> Path:
    return bundled_models_dir() / "trocr" / model_name.replace("/", "--")


def trocr_model_cache_dir(model_name: str) -> Path:
    return trocr_cache_root() / "hub" / f"models--{model_name.replace('/', '--')}"


def bundled_trocr_model_cache_dir(model_name: str) -> Path:
    return bundled_models_dir() / "huggingface" / "hub" / f"models--{model_name.replace('/', '--')}"


def ensure_trocr_model_available(model_name: str) -> None:
    target_model_dir = trocr_model_dir(model_name)
    if target_model_dir.exists():
        return
    bundled_model_dir = bundled_trocr_model_dir(model_name)
    if bundled_model_dir.exists():
        target_model_dir.parent.mkdir(parents=True, exist_ok=True)
        shutil.copytree(bundled_model_dir, target_model_dir, dirs_exist_ok=True)
        return

    target_dir = trocr_model_cache_dir(model_name)
    if target_dir.exists():
        return
    bundled_dir = bundled_trocr_model_cache_dir(model_name)
    if bundled_dir.exists():
        target_dir.parent.mkdir(parents=True, exist_ok=True)
        shutil.copytree(bundled_dir, target_dir, dirs_exist_ok=True)


def resolve_trocr_model_source(model_name: str) -> str:
    local_dir = trocr_model_dir(model_name)
    if local_dir.exists():
        return str(local_dir)
    bundled_dir = bundled_trocr_model_dir(model_name)
    if bundled_dir.exists():
        return str(bundled_dir)
    return model_name


def find_model_dir(model_name: str) -> Optional[Path]:
    for base_dir in resolve_model_dirs():
        candidate = base_dir / model_name
        if candidate.exists():
            return candidate
    return None


def module_available(name: str) -> bool:
    return importlib.util.find_spec(name) is not None


def compute_file_sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def get_runtime_report(
    ocr_backend: str = "normal",
    handwriting_backend: str = "none",
) -> dict[str, Any]:
    paddle_models = [
        "PP-LCNet_x1_0_doc_ori",
        "UVDoc",
        "PP-LCNet_x1_0_textline_ori",
        "PP-OCRv5_server_det",
        "en_PP-OCRv5_mobile_rec",
    ]
    trocr_modules = ["torch", "transformers", "sentencepiece"]
    report = {
        "python_version": f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
        "ocr_backend": ocr_backend,
        "handwriting_backend": handwriting_backend,
        "normal_backend_ready": module_available("cv2")
        and module_available("fitz")
        and module_available("numpy")
        and module_available("openpyxl")
        and module_available("pytesseract")
        and module_available("rapidocr_onnxruntime"),
        "ai_backend_supported_python": sys.version_info < (3, 12),
        "ai_backend_dependencies": {
            "paddleocr": module_available("paddleocr"),
            "paddlepaddle": module_available("paddle"),
        },
        "ai_backend_models": {
            model_name: str(find_model_dir(model_name) or "")
            for model_name in paddle_models
        },
        "trocr_dependencies": {
            module_name: module_available(module_name)
            for module_name in trocr_modules
        },
    }
    report["missing_ai_dependencies"] = [
        name for name, available in report["ai_backend_dependencies"].items()
        if not available
    ]
    report["missing_ai_models"] = [
        name for name, path in report["ai_backend_models"].items()
        if not path
    ]
    report["missing_trocr_dependencies"] = [
        name for name, available in report["trocr_dependencies"].items()
        if not available
    ]
    report["ai_backend_ready"] = (
        report["ai_backend_supported_python"]
        and all(report["ai_backend_dependencies"].values())
    )
    report["trocr_ready"] = all(report["trocr_dependencies"].values())
    return report


def validate_runtime_requirements(
    ocr_backend: str,
    handwriting_backend: str = "none",
) -> None:
    report = get_runtime_report(
        ocr_backend=ocr_backend,
        handwriting_backend=handwriting_backend,
    )
    if not report["normal_backend_ready"]:
        raise RuntimeError(
            "Normal OCR dependencies are missing. Install requirements-normal.txt."
        )
    if ocr_backend == "ai":
        if not report["ai_backend_supported_python"]:
            raise RuntimeError(
                "AI OCR requires Python 3.10 or 3.11 in this project."
            )
        missing = [
            name for name, available in report["ai_backend_dependencies"].items()
            if not available
        ]
        if missing:
            raise RuntimeError(
                "AI OCR dependencies are missing: "
                + ", ".join(missing)
                + ". Use setup_ai_env.sh or setup_ai_env.bat."
            )
    if handwriting_backend == "trocr":
        missing = [
            name for name, available in report["trocr_dependencies"].items()
            if not available
        ]
        if missing:
            raise RuntimeError(
                "TrOCR dependencies are missing: "
                + ", ".join(missing)
                + ". Use setup_ai_env.sh or setup_ai_env.bat."
            )


def configure_model_cache() -> None:
    cache_home = user_data_root() / "models"
    cache_home.mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("PADDLE_PDX_CACHE_HOME", str(cache_home))
    hf_home = trocr_cache_root()
    hf_home.mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("HF_HOME", str(hf_home))
    os.environ.setdefault("TRANSFORMERS_CACHE", str(hf_home / "hub"))


def detect_tesseract_cmd() -> Optional[str]:
    candidates = [
        shutil.which("tesseract"),
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)
    return None


def validate_runtime(ocr_backend: str) -> None:
    if ocr_backend == "ai" and sys.version_info >= (3, 12):
        raise RuntimeError(
            "PaddleOCR packaging is not available in this project on Python "
            f"{sys.version_info.major}.{sys.version_info.minor}. "
            "Use Python 3.10 or 3.11 for --ocr-backend ai, or run with "
            "--ocr-backend normal on newer Python versions."
        )


def render_pdf_to_images(pdf_path: Path, dpi: int = 300) -> List[np.ndarray]:
    zoom = dpi / 72.0
    matrix = fitz.Matrix(zoom, zoom)
    images: List[np.ndarray] = []

    with fitz.open(pdf_path) as pdf:
        for page in pdf:
            pix = page.get_pixmap(matrix=matrix, alpha=False)
            image = np.frombuffer(pix.samples, dtype=np.uint8).reshape(
                pix.height, pix.width, pix.n
            )
            if pix.n == 4:
                image = cv2.cvtColor(image, cv2.COLOR_BGRA2BGR)
            else:
                image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)
            images.append(image)

    return images


def preprocess_for_detection(image: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    thresh = cv2.threshold(
        blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU
    )[1]
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (9, 9))
    return cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=2)


def merge_boxes(boxes: Sequence[BoundingBox], gap: int = 25) -> List[BoundingBox]:
    merged: List[BoundingBox] = []
    for box in sorted(boxes, key=lambda item: (item[1], item[0])):
        x, y, w, h = box
        expanded = (x - gap, y - gap, w + 2 * gap, h + 2 * gap)
        merged_this_round = False
        for index, current in enumerate(merged):
            cx, cy, cw, ch = current
            if not (
                expanded[0] > cx + cw
                or expanded[0] + expanded[2] < cx
                or expanded[1] > cy + ch
                or expanded[1] + expanded[3] < cy
            ):
                nx = min(x, cx)
                ny = min(y, cy)
                nw = max(x + w, cx + cw) - nx
                nh = max(y + h, cy + ch) - ny
                merged[index] = (nx, ny, nw, nh)
                merged_this_round = True
                break
        if not merged_this_round:
            merged.append(box)
    return merged


def detect_document_regions(
    image: np.ndarray, min_area_ratio: float = 0.08
) -> List[BoundingBox]:
    processed = preprocess_for_detection(image)
    contours, _ = cv2.findContours(
        processed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
    )
    image_area = image.shape[0] * image.shape[1]
    min_area = int(image_area * min_area_ratio)
    candidates: List[BoundingBox] = []

    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        area = w * h
        if area < min_area:
            continue
        if w < 150 or h < 150:
            continue
        candidates.append((x, y, w, h))

    merged = merge_boxes(candidates)
    if not merged:
        return [(0, 0, image.shape[1], image.shape[0])]
    return sorted(merged, key=lambda item: (item[1], item[0]))


def crop_with_padding(
    image: np.ndarray, box: BoundingBox, padding: int = 20
) -> np.ndarray:
    x, y, w, h = box
    height, width = image.shape[:2]
    x1 = max(0, x - padding)
    y1 = max(0, y - padding)
    x2 = min(width, x + w + padding)
    y2 = min(height, y + h + padding)
    return image[y1:y2, x1:x2]


def crop_relative(
    image: np.ndarray,
    x1_ratio: float,
    y1_ratio: float,
    x2_ratio: float,
    y2_ratio: float,
) -> np.ndarray:
    height, width = image.shape[:2]
    x1 = max(0, min(width, int(width * x1_ratio)))
    y1 = max(0, min(height, int(height * y1_ratio)))
    x2 = max(x1 + 1, min(width, int(width * x2_ratio)))
    y2 = max(y1 + 1, min(height, int(height * y2_ratio)))
    return image[y1:y2, x1:x2]


def preprocess_for_ocr(image: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    normalized = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)
    clahe = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8, 8))
    contrast = clahe.apply(normalized)
    denoised = cv2.fastNlMeansDenoising(contrast, None, 18, 7, 21)
    sharpened = cv2.addWeighted(denoised, 1.4, cv2.GaussianBlur(denoised, (0, 0), 3), -0.4, 0)
    binary = cv2.adaptiveThreshold(
        sharpened,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31,
        11,
    )
    return deskew_binary_image(binary)


def preprocess_for_tesseract(image: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    normalized = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    contrast = clahe.apply(normalized)
    denoised = cv2.fastNlMeansDenoising(contrast, None, 12, 7, 21)
    binary = cv2.adaptiveThreshold(
        denoised,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31,
        9,
    )
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    cleaned = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel)
    return deskew_binary_image(cleaned)


def preprocess_for_tesseract_strong(image: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    upscaled = cv2.resize(gray, None, fx=1.8, fy=1.8, interpolation=cv2.INTER_CUBIC)
    normalized = cv2.normalize(upscaled, None, 0, 255, cv2.NORM_MINMAX)
    clahe = cv2.createCLAHE(clipLimit=3.5, tileGridSize=(8, 8))
    contrast = clahe.apply(normalized)
    denoised = cv2.fastNlMeansDenoising(contrast, None, 18, 7, 21)
    blur = cv2.medianBlur(denoised, 3)
    binary = cv2.adaptiveThreshold(
        blur,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        41,
        7,
    )
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    cleaned = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
    return deskew_binary_image(cleaned)


def preprocess_for_ocr_detail(image: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    upscaled = cv2.resize(gray, None, fx=2.0, fy=2.0, interpolation=cv2.INTER_CUBIC)
    clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8, 8))
    contrast = clahe.apply(upscaled)
    sharpened = cv2.addWeighted(
        contrast,
        1.6,
        cv2.GaussianBlur(contrast, (0, 0), 2),
        -0.6,
        0,
    )
    binary = cv2.adaptiveThreshold(
        sharpened,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        41,
        9,
    )
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    cleaned = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
    return deskew_binary_image(cleaned)


def deskew_binary_image(binary: np.ndarray) -> np.ndarray:
    coords = np.column_stack(np.where(binary < 250))
    if coords.size == 0:
        return binary
    angle = cv2.minAreaRect(coords)[-1]
    if angle < -45:
        angle += 90
    if -0.5 < angle < 0.5:
        return binary
    height, width = binary.shape[:2]
    matrix = cv2.getRotationMatrix2D((width // 2, height // 2), angle, 1.0)
    return cv2.warpAffine(
        binary,
        matrix,
        (width, height),
        flags=cv2.INTER_CUBIC,
        borderMode=cv2.BORDER_REPLICATE,
    )


def normalize_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def contains_arabic(value: str) -> bool:
    return any("\u0600" <= char <= "\u06FF" for char in value)


def contains_latin(value: str) -> bool:
    return any(("A" <= char <= "Z") or ("a" <= char <= "z") for char in value)


def alpha_ratio(value: str) -> float:
    if not value:
        return 0.0
    alnum = sum(char.isalnum() for char in value)
    alpha = sum(char.isalpha() for char in value)
    if alnum == 0:
        return 0.0
    return alpha / alnum


def sanitize_filename_part(value: str) -> str:
    value = normalize_whitespace(value)
    value = re.sub(r"[\\/:*?\"<>|]+", "", value)
    value = re.sub(r"\s+", "", value)
    return value or "Unknown"


def sanitize_filename(value: str) -> str:
    value = normalize_whitespace(value)
    value = re.sub(r"[\\/:*?\"<>|]+", "", value)
    return value.strip(" .") or "Unknown"


def normalize_number_candidate(value: str) -> str:
    value = normalize_whitespace(value).strip(" -:/._")
    value = re.sub(r"[^A-Za-z0-9\-\/]", "", value)
    prefix_match = re.match(r"^([A-Za-z]{2,}[-\/]?)(.*)$", value)
    prefix = ""
    suffix = value
    if prefix_match and any(char.isdigit() for char in prefix_match.group(2)):
        prefix = prefix_match.group(1)
        suffix = prefix_match.group(2)
    substitutions = str.maketrans(
        {
            "O": "0",
            "o": "0",
            "I": "1",
            "l": "1",
            "S": "5",
            "B": "8",
            "Z": "2",
        }
    )
    suffix = suffix.translate(substitutions)
    value = prefix + suffix
    return value or "Unknown"


def is_plausible_number(value: str) -> bool:
    normalized = normalize_number_candidate(value)
    digit_count = sum(char.isdigit() for char in normalized)
    if digit_count < 3:
        return False
    if re.fullmatch(r"[A-Za-z]+", normalized):
        return False
    return True


def normalize_ocr_text(text: str) -> str:
    replacements = (
        (r"(?i)-?a(?:d|u)n[o0][cg]\s+distribut(?:ion|lan|lon|inn|l0n|i0n)\b", "ADNOC Distribution"),
        (r"(?i)\binv[o0]ice\b", "invoice"),
        (r"(?i)\biny[o0]ice\b", "invoice"),
        (r"(?i)\btax\s+iny[o0]ice\b", "tax invoice"),
        (r"(?i)\brece[il1]pt\b", "receipt"),
        (r"(?i)\banmount\b", "amount"),
        (r"(?i)\bmernbership\b", "membership"),
        (r"(?i)\battendant\s+1d\b", "Attendant ID"),
        (r"(?i)\bt1d\b", "TID"),
        (r"(?i)\byat\b", "VAT"),
        (r"(?i)\badno[c0]\b", "ADNOC"),
        (r"(?i)\badnog\b", "ADNOC"),
        (r"(?i)\baunoc\b", "ADNOC"),
        (r"(?i)\badnc\b", "ADNOC"),
    )
    normalized_lines: List[str] = []
    for raw_line in text.splitlines():
        line = normalize_whitespace(raw_line)
        if not line:
            continue
        for pattern, replacement in replacements:
            line = re.sub(pattern, replacement, line)
        normalized_lines.append(line)
    return "\n".join(normalized_lines)


def text_lines(text: str) -> List[str]:
    return [normalize_whitespace(line) for line in text.splitlines() if normalize_whitespace(line)]


def canonical_letters(value: str) -> str:
    return re.sub(r"[^a-z]", "", value.lower())


def looks_like_document_label(value: str) -> bool:
    canonical = canonical_letters(value)
    if not canonical:
        return False
    if any(token in canonical for token in ("taxinvoice", "invoice", "receipt", "cashsale", "document")):
        return True
    for target, threshold in (("taxinvoice", 0.7), ("invoice", 0.78), ("receipt", 0.78)):
        if SequenceMatcher(None, canonical, target).ratio() >= threshold:
            return True
    return False


def extract_number_after_labels(
    lines: Sequence[str],
    label_patterns: Sequence[str],
    max_lookahead: int = 1,
) -> List[str]:
    found: List[str] = []
    candidate_pattern = r"[A-Z0-9OISBZl][A-Z0-9OISBZl\-\/]{2,}"

    for index, line in enumerate(lines):
        lowered = line.lower()
        match_end: Optional[int] = None
        for pattern in label_patterns:
            match = re.search(pattern, lowered)
            if match:
                match_end = match.end()
                break
        if match_end is None:
            continue

        search_space = [line[match_end:]]
        for offset in range(1, max_lookahead + 1):
            if index + offset < len(lines):
                search_space.append(lines[index + offset])

        for candidate_line in search_space:
            for candidate in re.findall(candidate_pattern, candidate_line.upper()):
                value = normalize_number_candidate(candidate)
                if not is_plausible_number(value):
                    continue
                found.append(value)
    return found


def best_number(candidates: Sequence[str]) -> str:
    filtered = []
    for candidate in candidates:
        normalized = normalize_number_candidate(candidate)
        digit_count = sum(char.isdigit() for char in normalized)
        if not is_plausible_number(normalized):
            continue
        filtered.append((digit_count, len(normalized), normalized))
    if not filtered:
        return "Unknown"
    filtered.sort(reverse=True)
    return filtered[0][2]


def choose_number_by_type(text: str, doc_type: str) -> str:
    lines = text_lines(text)

    invoice_labeled = extract_number_after_labels(
        lines,
        [r"\binvoice\s*(?:no|number|#)\b", r"\btax\s+invoice\b"],
        max_lookahead=1,
    )
    receipt_labeled = extract_number_after_labels(
        lines,
        [r"\breceipt\s*(?:no|number|#)\b"],
        max_lookahead=1,
    )
    token_labeled = extract_number_after_labels(
        lines,
        [r"\btoken\s+number\b", r"\bapproval\s+code\b"],
        max_lookahead=0,
    )
    banned_values = {normalize_number_candidate(value) for value in token_labeled}

    invoice_patterns = [
        r"(?i)\binvoice\s*(?:no|#|number)\s*[:;\-]?\s*([A-Z0-9OISBZl\-\/]{4,})",
        r"(?i)\btax\s+invoice\b.*?\b(?:no|#|number)\b\s*[:;\-]?\s*([A-Z0-9OISBZl\-\/]{4,})",
    ]
    receipt_patterns = [
        r"(?i)\breceipt\s*(?:no|#|number)?\s*[:;\-]?\s*([A-Z0-9OISBZl\-\/]{4,})",
    ]
    generic_patterns = [
        r"(?i)\b(?:bill|voucher|document|inv|trx|ref|serial)\s*(?:no|#|number)?\s*[:;\-]?\s*([A-Z0-9OISBZl\-\/]{3,})",
        r"(?i)\b(?:no|number)\b\s*[:;\-]?\s*([A-Z0-9OISBZl\-\/]{3,})",
    ]

    def collect(patterns: Sequence[str]) -> List[str]:
        values: List[str] = []
        for line in lines:
            for pattern in patterns:
                for match in re.finditer(pattern, line):
                    value = normalize_number_candidate(match.group(1))
                    if is_plausible_number(value) and value not in banned_values:
                        values.append(value)
        return values

    invoice_values = [value for value in invoice_labeled if value not in banned_values] + collect(invoice_patterns)
    receipt_values = [value for value in receipt_labeled if value not in banned_values] + collect(receipt_patterns)
    generic_values = collect(generic_patterns)

    if doc_type == "Invoice" and invoice_values:
        return best_number(invoice_values)
    if doc_type == "Receipt" and receipt_values:
        return best_number(receipt_values)
    if invoice_values:
        return best_number(invoice_values)
    if receipt_values:
        return best_number(receipt_values)

    for index, line in enumerate(lines):
        if re.search(r"(?i)\b(?:invoice|receipt|bill|voucher|document|inv|trx|ref)\b", line):
            for offset in range(0, 3):
                if index + offset >= len(lines):
                    continue
                candidates = re.findall(r"[A-Z0-9OISBZl][A-Z0-9OISBZl\-\/]{2,}", lines[index + offset])
                for candidate in candidates:
                    value = normalize_number_candidate(candidate)
                    if is_plausible_number(value) and value not in banned_values:
                        return value

    if generic_values:
        return best_number([value for value in generic_values if value not in banned_values])
    return "Unknown"


def make_ocr_engine(
    language: str = "en",
    use_angle_cls: bool = True,
    text_recognition_model_name: Optional[str] = None,
) -> Any:
    try:
        from paddleocr import PaddleOCR
    except ImportError as exc:
        raise RuntimeError(
            "PaddleOCR is not installed. Run 'pip install paddleocr paddlepaddle'."
        ) from exc
    os.environ.setdefault("DISABLE_MODEL_SOURCE_CHECK", "True")
    os.environ.setdefault("PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK", "True")
    kwargs = {
        "use_textline_orientation": use_angle_cls,
    }
    if text_recognition_model_name:
        kwargs["text_recognition_model_name"] = text_recognition_model_name
    else:
        kwargs["lang"] = language
    return PaddleOCR(**kwargs)


def make_ocr_engines(
    language: str = "en",
    use_angle_cls: bool = True,
    ocr_profile: str = "mixed",
) -> List[Any]:
    if ocr_profile == "printed":
        return [make_ocr_engine(language=language, use_angle_cls=use_angle_cls)]
    if ocr_profile == "handwriting":
        return [
            make_ocr_engine(
                language=language,
                use_angle_cls=use_angle_cls,
                text_recognition_model_name="PP-OCRv5_server_rec",
            )
        ]
    return [
        make_ocr_engine(language=language, use_angle_cls=use_angle_cls),
        make_ocr_engine(
            language=language,
            use_angle_cls=use_angle_cls,
            text_recognition_model_name="PP-OCRv5_server_rec",
        ),
    ]


def make_trocr_engine(model_name: str = "microsoft/trocr-base-handwritten") -> Any:
    try:
        import torch
        from transformers import TrOCRProcessor, VisionEncoderDecoderModel
    except ImportError as exc:
        raise RuntimeError(
            "TrOCR dependencies are not installed. Run "
            "'pip install transformers torch sentencepiece'."
        ) from exc

    if torch.backends.mps.is_available():
        device = "mps"
    elif torch.cuda.is_available():
        device = "cuda"
    else:
        device = "cpu"

    ensure_trocr_model_available(model_name)
    configure_model_cache()
    model_source = resolve_trocr_model_source(model_name)
    processor = TrOCRProcessor.from_pretrained(model_source)
    model = VisionEncoderDecoderModel.from_pretrained(
        model_source,
        use_safetensors=True,
    )
    model.to(device)
    model.eval()
    return {
        "processor": processor,
        "model": model,
        "device": device,
        "torch": torch,
    }


def extract_text_with_trocr(image: np.ndarray, trocr_engine: Any) -> str:
    processor = trocr_engine["processor"]
    model = trocr_engine["model"]
    device = trocr_engine["device"]
    torch = trocr_engine["torch"]

    rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    pil_image = Image.fromarray(rgb)

    with torch.no_grad():
        pixel_values = processor(images=pil_image, return_tensors="pt").pixel_values
        pixel_values = pixel_values.to(device)
        generated_ids = model.generate(pixel_values)
        text = processor.batch_decode(generated_ids, skip_special_tokens=True)[0]

    return normalize_whitespace(text)


def extract_text_with_tesseract(image: np.ndarray) -> Tuple[str, np.ndarray]:
    variants = [
        cv2.cvtColor(image, cv2.COLOR_BGR2RGB),
        cv2.cvtColor(preprocess_for_tesseract(image), cv2.COLOR_GRAY2RGB),
        cv2.cvtColor(preprocess_for_tesseract_strong(image), cv2.COLOR_GRAY2RGB),
    ]
    best_text = ""
    best_variant = variants[0]
    best_score = -1
    config = "--oem 3 --psm 6"
    for variant in variants:
        text = pytesseract.image_to_string(variant, config=config)
        lines = [normalize_whitespace(line) for line in text.splitlines() if normalize_whitespace(line)]
        score = score_text(lines)
        if score > best_score:
            best_text = "\n".join(lines)
            best_variant = variant
            best_score = score
    return best_text, best_variant


def extract_text_with_rapidocr(image: np.ndarray) -> Tuple[str, np.ndarray]:
    try:
        from rapidocr_onnxruntime import RapidOCR
    except ImportError as exc:
        raise RuntimeError(
            "RapidOCR is not installed. Run 'pip install rapidocr-onnxruntime'."
        ) from exc
    variants = [
        image,
        cv2.cvtColor(preprocess_for_tesseract(image), cv2.COLOR_GRAY2BGR),
        cv2.cvtColor(preprocess_for_tesseract_strong(image), cv2.COLOR_GRAY2BGR),
        cv2.cvtColor(preprocess_for_ocr_detail(image), cv2.COLOR_GRAY2BGR),
    ]
    engine = RapidOCR()
    best_text = ""
    best_variant = variants[0]
    best_score = -1
    for variant in variants:
        result, _ = engine(variant)
        lines: List[str] = []
        if result:
            for item in result:
                if len(item) >= 2:
                    cleaned = normalize_whitespace(str(item[1]))
                    if cleaned:
                        lines.append(cleaned)
        score = score_text(lines)
        if score > best_score:
            best_text = "\n".join(lines)
            best_variant = variant
            best_score = score
    return best_text, best_variant


def flatten_paddle_result(result: object) -> List[str]:
    texts: List[str] = []
    if result is None:
        return texts
    if isinstance(result, str):
        cleaned = normalize_whitespace(result)
        if cleaned:
            texts.append(cleaned)
        return texts
    if isinstance(result, (list, tuple)):
        if len(result) == 2 and isinstance(result[0], str):
            cleaned = normalize_whitespace(result[0])
            if cleaned:
                texts.append(cleaned)
            return texts
        for item in result:
            texts.extend(flatten_paddle_result(item))
    return texts


def result_to_lines(result: Any) -> List[str]:
    lines: List[str] = []
    for item in result:
        if hasattr(item, "get"):
            rec_texts = item.get("rec_texts", [])
            for text in rec_texts:
                cleaned = normalize_whitespace(str(text))
                if cleaned:
                    lines.append(cleaned)
        else:
            lines.extend(flatten_paddle_result(item))
    return lines


def score_text(lines: Sequence[str]) -> int:
    text = "\n".join(lines)
    digit_bonus = sum(char.isdigit() for char in text) * 2
    alpha_bonus = sum(char.isalpha() for char in text)
    return len(text) + digit_bonus + alpha_bonus


def extract_text(image: np.ndarray, ocr_engines: Sequence[Any]) -> Tuple[str, np.ndarray]:
    variants = [
        cv2.cvtColor(image, cv2.COLOR_BGR2RGB),
        cv2.cvtColor(preprocess_for_ocr(image), cv2.COLOR_GRAY2RGB),
        cv2.cvtColor(preprocess_for_tesseract_strong(image), cv2.COLOR_GRAY2RGB),
        cv2.cvtColor(preprocess_for_ocr_detail(image), cv2.COLOR_GRAY2RGB),
    ]
    best_lines: List[str] = []
    best_variant = variants[0]
    best_score = -1

    for ocr_engine in ocr_engines:
        for variant in variants:
            result = ocr_engine.predict(variant)
            lines = result_to_lines(result)
            score = score_text(lines)
            if score > best_score:
                best_lines = lines
                best_variant = variant
                best_score = score

    return "\n".join(best_lines), best_variant


def extract_text_by_backend(
    image: np.ndarray, ocr_backend: str, ocr_engines: Sequence[Any]
) -> Tuple[str, np.ndarray]:
    if ocr_backend == "normal":
        configured_cmd = getattr(pytesseract.pytesseract, "tesseract_cmd", "") or ""
        tesseract_cmd = configured_cmd if configured_cmd and Path(configured_cmd).exists() else detect_tesseract_cmd()
        if tesseract_cmd:
            pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
            try:
                return extract_text_with_tesseract(image)
            except Exception:
                pass
        return extract_text_with_rapidocr(image)
    return extract_text(image, ocr_engines)


def infer_doc_type(text: str) -> str:
    lowered = text.lower()
    if "invoice" in lowered or "tax invoice" in lowered:
        return "Invoice"
    if "receipt" in lowered or "cash sale" in lowered:
        return "Receipt"
    if "document" in lowered:
        return "Document"
    for line in text_lines(text):
        if looks_like_document_label(line):
            canonical = canonical_letters(line)
            if "receipt" in canonical:
                return "Receipt"
            return "Invoice"
    return "Document"


def parse_date(text: str) -> str:
    lines = text_lines(text)
    for index, line in enumerate(lines):
        if "date" not in line.lower():
            continue
        search_space = [line]
        if index + 1 < len(lines):
            search_space.append(lines[index + 1])
        for candidate in search_space:
            for pattern in (
                r"\b(\d{4}[/-]\d{1,2}[/-]\d{1,2})\b",
                r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b",
                r"\b(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})\b",
                r"\b([A-Za-z]{3,9}\s+\d{1,2},\s*\d{4})\b",
            ):
                match = re.search(pattern, candidate, flags=re.IGNORECASE)
                if match:
                    raw = normalize_whitespace(match.group(1).replace(".", "/"))
                    for fmt in (
                        "%Y-%m-%d",
                        "%Y/%m/%d",
                        "%d-%m-%Y",
                        "%d/%m/%Y",
                        "%m-%d-%Y",
                        "%m/%d/%Y",
                        "%d-%m-%y",
                        "%d/%m/%y",
                        "%d %b %Y",
                        "%d %B %Y",
                        "%d %b %y",
                        "%d %B %y",
                        "%b %d, %Y",
                        "%B %d, %Y",
                    ):
                        try:
                            parsed = datetime.strptime(raw, fmt)
                            return parsed.strftime("%Y-%m-%d")
                        except ValueError:
                            continue

    patterns = [
        r"\b(\d{4}[/-]\d{1,2}[/-]\d{1,2})\b",
        r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b",
        r"\b(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})\b",
        r"\b([A-Za-z]{3,9}\s+\d{1,2},\s*\d{4})\b",
    ]
    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%m-%d-%Y",
        "%m/%d/%Y",
        "%d-%m-%y",
        "%d/%m/%y",
        "%d %b %Y",
        "%d %B %Y",
        "%d %b %y",
        "%d %B %y",
        "%b %d, %Y",
        "%B %d, %Y",
    ]

    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            raw = normalize_whitespace(match.group(1).replace(".", "/"))
            for fmt in formats:
                try:
                    parsed = datetime.strptime(raw, fmt)
                    return parsed.strftime("%Y-%m-%d")
                except ValueError:
                    continue
    return "Unknown"


def parse_number(text: str) -> str:
    doc_type = infer_doc_type(text)
    return choose_number_by_type(text, doc_type)


def parse_amount(text: str) -> str:
    lines = text_lines(text)

    def line_values(line: str) -> List[str]:
        found: List[str] = []
        for token in re.findall(r"\d[\d,]*(?:\.\d+)+|\d[\d,]*", line):
            cleaned = token.replace(",", "")
            if cleaned.count(".") <= 1:
                if re.fullmatch(r"\d+\.\d{1,2}", cleaned):
                    found.append(cleaned)
                continue
            parts = cleaned.split(".")
            for index in range(len(parts) - 1):
                left = parts[index]
                right = parts[index + 1]
                if not left or not right:
                    continue
                if 1 <= len(right) <= 2:
                    found.append(f"{left}.{right}")
                    if len(left) > 4:
                        found.append(f"{left[-3:]}.{right}")
                        found.append(f"{left[-2:]}.{right}")
        for left, right in re.findall(r"\b(\d{1,5})\s+(\d{2})\b", line):
            found.append(f"{left}.{right}")
        return found

    scores: dict[str, int] = {}
    counts: dict[str, int] = {}

    for index, line in enumerate(lines):
        values = line_values(line)
        lowered = line.lower()
        line_bonus = 0
        if re.search(r"(?i)\bgrand\s+total\b|\btotal\s+amount\b|\btotal\s+aed\b|\btotal\b|\bamount\s+due\b", line):
            line_bonus += 100
        if re.search(r"(?i)\bpayment\b|\bpaid\b|\bmembership\b", line):
            line_bonus += 60
        if re.search(r"(?i)\bsubtotal\b|\bsub\s*total\b", line):
            line_bonus += 30
        if re.search(r"(?i)\bvat\b|\btax\b", line):
            line_bonus -= 15
        if re.search(r"(?i)\bqty\b|\bprice\b", line):
            line_bonus -= 10
        if re.search(r"\d+\.\d{2}\s*[-:]\s*\d{3,}", line):
            line_bonus -= 160
        if len(re.findall(r"\d+\.\d{2}", line)) >= 2 and not re.search(
            r"(?i)\bgrand\s+total\b|\btotal\s+amount\b|\btotal\s+aed\b|\btotal\b|\bamount\s+due\b",
            line,
        ):
            line_bonus -= 120

        if not values and index + 1 < len(lines):
            next_values = line_values(lines[index + 1])
            if next_values and re.search(
                r"(?i)\bgrand\s+total\b|\btotal\s+amount\b|\btotal\s+aed\b|\btotal\b|\bamount\s+due\b",
                line,
            ):
                values = next_values
                line_bonus += 120

        for value in values:
            try:
                amount = float(value)
            except ValueError:
                continue
            if amount <= 0:
                continue
            normalized = f"{amount:.2f}"
            counts[normalized] = counts.get(normalized, 0) + 1
            scores[normalized] = scores.get(normalized, 0) + line_bonus

    if scores:
        ranked = []
        for value, score in scores.items():
            amount = float(value)
            frequency_bonus = (counts.get(value, 1) - 1) * 60
            ranked.append((score + frequency_bonus, counts.get(value, 1), amount, value))
        ranked.sort(key=lambda item: (item[0], item[1], item[2]), reverse=True)
        return ranked[0][3]
    return "Unknown"


def parse_company_name(text: str) -> str:
    blacklist = {
        "invoice",
        "tax invoice",
        "receipt",
        "document",
        "bill to",
        "ship to",
        "customer copy",
        "original",
        "duplicate",
        "thank you",
        "contact us",
        "all kinds of building materials",
        "wholesale & retails",
        "direct order",
        "cashier",
        "pos",
    }
    preferred_terms = {
        "distribution",
        "trading",
        "company",
        "co",
        "corp",
        "corporation",
        "llc",
        "l.l.c",
        "limited",
        "ltd",
        "services",
        "station",
        "building",
        "materials",
        "transport",
        "industries",
        "group",
        "market",
        "store",
        "shop",
        "adnoc",
    }
    item_terms = {
        "chicken",
        "mandi",
        "rice",
        "quarter",
        "half",
        "full",
        "burger",
        "sandwich",
        "meal",
        "salad",
        "juice",
        "water",
        "coke",
        "pepsi",
        "tea",
        "coffee",
    }
    strong_company_patterns = (
        (r"(?i)\badnoc\s+distribution\b", "ADNOC Distribution"),
        (r"(?i)-?a(?:d|u)n[o0][cg]\s+distribut", "ADNOC Distribution"),
        (
            r"(?i)jesr\s+al\s+mad[il1]n[aia].*building.*materials",
            "JESR AL MADINA BUILDING MATERIALS TRAD. L.L.C.",
        ),
        (
            r"100342274600003",
            "JESR AL MADINA BUILDING MATERIALS TRAD. L.L.C.",
        ),
    )
    candidates: List[Tuple[int, str]] = []

    for pattern, company_name in strong_company_patterns:
        if re.search(pattern, text):
            return company_name

    for line_index, line in enumerate(text_lines(text)):
        if line_index > 7:
            continue
        cleaned = line
        if len(cleaned) < 3 or len(cleaned) > 60:
            continue
        lowered = cleaned.lower()
        if lowered in blacklist:
            continue
        if looks_like_document_label(cleaned):
            continue
        if re.search(r"\b(?:invoice|receipt|total|date|no\.?|number|vat|subtotal|price|qty|item|payment)\b", lowered):
            continue
        if re.search(r"\d{3,}", cleaned):
            continue
        if "po box" in lowered or "p.o box" in lowered or "abu dhabi" in lowered or "uae" in lowered:
            continue
        if "," in cleaned and not any(term in lowered for term in preferred_terms):
            continue
        if any(term in lowered for term in item_terms):
            continue
        if sum(char.isalpha() for char in cleaned) < 4:
            continue
        score = 0
        words = re.findall(r"[A-Za-z]+", cleaned.lower())
        has_preferred_term = any(word in preferred_terms for word in words)
        uppercase_letters = sum(char.isupper() for char in cleaned)
        lowercase_letters = sum(char.islower() for char in cleaned)
        if line_index > 4 and not has_preferred_term:
            continue
        if not has_preferred_term:
            if line_index > 2:
                continue
            if uppercase_letters < 4 or uppercase_letters < lowercase_letters:
                continue
        score += min(len(words), 6) * 3
        if contains_latin(cleaned) and not contains_arabic(cleaned):
            score += 25
        if contains_latin(cleaned):
            score += 10
        if alpha_ratio(cleaned) > 0.75:
            score += 8
        score += sum(12 for word in words if word in preferred_terms)
        if len(cleaned) >= 8:
            score += 5
        if line_index <= 3:
            score += 12
        if len(words) == 1 and len(words[0]) <= 5:
            score -= 20
        if "_" in cleaned:
            score -= 15
        if re.search(r"(?i)^[a-z]{2,4}_[a-z]{2,6}$", cleaned):
            score -= 25
        if re.search(r"(?i)^[a-z]{2,8}$", cleaned) and cleaned.lower() not in preferred_terms:
            score -= 18
        candidates.append((score, cleaned))

    if candidates:
        candidates.sort(key=lambda item: item[0], reverse=True)
        return candidates[0][1]
    return "Unknown"


def choose_best_date_candidate(texts: Sequence[str]) -> str:
    for text in texts:
        value = parse_date(normalize_ocr_text(text))
        if value != "Unknown":
            return value
    return "Unknown"


def choose_best_number_candidate(texts: Sequence[str]) -> str:
    for text in texts:
        value = parse_number(normalize_ocr_text(text))
        if value != "Unknown":
            return value
    return "Unknown"


def choose_best_amount_candidate(texts: Sequence[str]) -> str:
    for text in texts:
        value = parse_amount(normalize_ocr_text(text))
        if value == "Unknown":
            cleaned = normalize_whitespace(text)
            integer_tokens = re.findall(r"\b\d{1,5}\b", cleaned)
            if len(integer_tokens) == 1:
                try:
                    amount = float(integer_tokens[0])
                    if amount >= 1.0:
                        return f"{amount:.2f}"
                except ValueError:
                    pass
            continue
        try:
            if float(value) >= 1.0:
                return value
        except ValueError:
            continue
    return "Unknown"


def extract_fields(text: str) -> DocumentFields:
    normalized_text = normalize_ocr_text(text)
    return DocumentFields(
        doc_type=infer_doc_type(normalized_text),
        date=parse_date(normalized_text),
        number=parse_number(normalized_text),
        company_name=parse_company_name(normalized_text),
        amount=parse_amount(normalized_text),
    )


def compute_confidence(fields: DocumentFields) -> tuple[int, str]:
    score = 0

    if fields.doc_type != "Unknown":
        score += 20
    if fields.date != "Unknown" and re.fullmatch(r"\d{4}-\d{2}-\d{2}", fields.date):
        score += 20
    if fields.number != "Unknown" and len(fields.number) >= 3:
        score += 20
    if fields.company_name != "Unknown" and len(fields.company_name) >= 4:
        score += 20
    if fields.amount != "Unknown":
        try:
            amount_value = float(fields.amount)
            if amount_value > 0:
                score += 20
        except ValueError:
            pass

    if fields.company_name == "Unknown" or len(fields.company_name) < 4:
        score -= 8
    if fields.number != "Unknown" and re.search(r"[^A-Za-z0-9\-/]", fields.number):
        score -= 6

    score = max(0, min(100, score))
    if score >= 80:
        label = "high"
    elif score >= 50:
        label = "medium"
    else:
        label = "low"
    return score, label


def refine_fields_with_region_ocr(
    image: np.ndarray,
    fields: DocumentFields,
    ocr_backend: str,
    ocr_engines: Sequence[Any],
    trocr_engine: Any = None,
) -> DocumentFields:
    needs_number = fields.number == "Unknown"
    needs_date = fields.date == "Unknown"
    needs_amount = fields.amount == "Unknown"
    needs_company = fields.company_name == "Unknown"

    if not any((needs_number, needs_date, needs_amount, needs_company)):
        return fields

    region_specs = {
        "header_left": (0.0, 0.0, 0.6, 0.32),
        "header_right": (0.45, 0.0, 1.0, 0.32),
        "top_half": (0.0, 0.0, 1.0, 0.42),
        "bottom_right": (0.58, 0.62, 1.0, 1.0),
        "bottom_right_tight": (0.78, 0.86, 0.98, 0.99),
    }
    region_texts: dict[str, str] = {}

    for key, ratios in region_specs.items():
        region = crop_relative(image, *ratios)
        text, _ = extract_text_by_backend(region, ocr_backend, ocr_engines)
        region_texts[key] = normalize_ocr_text(text)

    if needs_number:
        number_candidate = parse_number(
            "\n".join(
                part
                for part in (
                    region_texts["header_left"],
                    region_texts["top_half"],
                    region_texts["header_right"],
                )
                if part
            )
        )
        if number_candidate != "Unknown":
            fields.number = number_candidate

    if needs_date:
        date_candidate = parse_date(
            "\n".join(
                part
                for part in (
                    region_texts["header_right"],
                    region_texts["top_half"],
                )
                if part
            )
        )
        if date_candidate != "Unknown":
            fields.date = date_candidate

    if needs_amount:
        amount_candidate = parse_amount(region_texts["bottom_right"])
        if amount_candidate != "Unknown":
            try:
                if float(amount_candidate) >= 1.0:
                    fields.amount = amount_candidate
            except ValueError:
                pass
        if fields.amount == "Unknown" and fields.doc_type == "Invoice":
            amount_candidate = choose_best_amount_candidate(
                [region_texts["bottom_right_tight"], region_texts["bottom_right"]]
            )
            if amount_candidate != "Unknown":
                fields.amount = amount_candidate

    if needs_company:
        company_candidate = parse_company_name(region_texts["top_half"])
        if company_candidate != "Unknown":
            fields.company_name = company_candidate

    if trocr_engine is None:
        return fields

    trocr_regions = {
        "number_box": crop_relative(image, 0.05, 0.16, 0.38, 0.28),
        "date_box": crop_relative(image, 0.64, 0.16, 0.98, 0.28),
        "amount_box": crop_relative(image, 0.76, 0.82, 0.99, 0.98),
    }
    trocr_texts: dict[str, str] = {}
    for key, region in trocr_regions.items():
        try:
            trocr_texts[key] = extract_text_with_trocr(region, trocr_engine)
        except Exception:
            trocr_texts[key] = ""

    if fields.number == "Unknown":
        number_candidate = choose_best_number_candidate(
            [trocr_texts["number_box"], trocr_texts["date_box"]]
        )
        if number_candidate != "Unknown":
            fields.number = number_candidate

    if fields.date == "Unknown":
        date_candidate = choose_best_date_candidate([trocr_texts["date_box"]])
        if date_candidate != "Unknown":
            fields.date = date_candidate

    if fields.amount == "Unknown":
        amount_candidate = choose_best_amount_candidate([trocr_texts["amount_box"]])
        if amount_candidate != "Unknown":
            fields.amount = amount_candidate

    return fields


def build_output_name(
    fields: DocumentFields,
    project_name: str,
    naming_pattern: str = DEFAULT_NAMING_PATTERN,
) -> str:
    amount_part = fields.amount
    if amount_part != "Unknown":
        amount_part = f"{amount_part}AED"
    values = {
        "doc_type": sanitize_filename_part(fields.doc_type),
        "date": sanitize_filename_part(fields.date),
        "number": sanitize_filename_part(fields.number),
        "company_name": sanitize_filename_part(fields.company_name),
        "amount": sanitize_filename_part(fields.amount),
        "amount_aed": sanitize_filename_part(amount_part),
        "project_name": sanitize_filename_part(project_name),
    }
    try:
        rendered = naming_pattern.format(**values)
    except KeyError as exc:
        raise ValueError(
            f"Unsupported naming pattern field: {exc.args[0]}"
        ) from exc
    rendered = sanitize_filename(rendered)
    if not rendered.lower().endswith(".pdf"):
        rendered += ".pdf"
    return rendered


def write_excel_summary(records: Sequence[ProcessedRecord], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Documents"
    headers = [
        "SourceFile",
        "SourcePath",
        "SourceHash",
        "SourceType",
        "SourceOrigin",
        "SourceTimestamp",
        "OutputFile",
        "Type",
        "Date",
        "Number",
        "CompanyName",
        "Amount",
        "Currency",
        "TransactionDirection",
        "ProjectName",
        "ConfidenceScore",
        "ConfidenceLabel",
        "MatchStatus",
        "MatchScore",
        "MatchedRecordSourceFile",
        "MatchedRecordOutputFile",
        "MatchedRecordSourceType",
        "MatchedRecordSourceTimestamp",
        "MatchedRecordDate",
        "MatchedRecordNumber",
        "MatchedRecordCompanyName",
        "MatchedRecordAmount",
        "MatchedRecordTransactionDirection",
        "MatchBasis",
        "RawText",
    ]
    sheet.append(headers)
    for record in records:
        sheet.append(
            [
                record.source_file,
                record.source_path,
                record.source_hash,
                record.source_type,
                record.source_origin,
                record.source_timestamp,
                record.output_file,
                record.doc_type,
                record.date,
                record.number,
                record.company_name,
                record.amount,
                record.currency,
                record.transaction_direction,
                record.project_name,
                record.confidence_score,
                record.confidence_label,
                record.match_status,
                record.match_score,
                record.matched_record_source_file,
                record.matched_record_output_file,
                record.matched_record_source_type,
                record.matched_record_source_timestamp,
                record.matched_record_date,
                record.matched_record_number,
                record.matched_record_company_name,
                record.matched_record_amount,
                record.matched_record_transaction_direction,
                record.match_basis,
                record.raw_text,
            ]
        )
    bank_sheet = workbook.create_sheet("Bank Reconciliation")
    bank_headers = [
        "Status",
        "MatchScore",
        "Date",
        "CompanyName",
        "Amount",
        "Currency",
        "TransactionDirection",
        "Number",
        "SourceFile",
        "SourceTimestamp",
        "MatchedRecordSourceFile",
        "MatchedRecordOutputFile",
        "MatchedRecordSourceType",
        "MatchedRecordSourceTimestamp",
        "MatchedRecordDate",
        "MatchedRecordNumber",
        "MatchedRecordCompanyName",
        "MatchedRecordAmount",
        "MatchedRecordTransactionDirection",
        "MatchBasis",
        "RawText",
    ]
    bank_sheet.append(bank_headers)
    fill_by_status = {
        "matched": PatternFill(fill_type="solid", fgColor="C6EFCE"),
        "missing_receipt": PatternFill(fill_type="solid", fgColor="FFC7CE"),
        "not_applicable": PatternFill(fill_type="solid", fgColor="D9D9D9"),
    }
    for record in records:
        if record.source_type != "sheet" or record.doc_type != "BankTransaction":
            continue
        bank_sheet.append(
            [
                record.match_status,
                record.match_score,
                record.date,
                record.company_name,
                record.amount,
                record.currency,
                record.transaction_direction,
                record.number,
                record.source_file,
                record.source_timestamp,
                record.matched_record_source_file,
                record.matched_record_output_file,
                record.matched_record_source_type,
                record.matched_record_source_timestamp,
                record.matched_record_date,
                record.matched_record_number,
                record.matched_record_company_name,
                record.matched_record_amount,
                record.matched_record_transaction_direction,
                record.match_basis,
                record.raw_text,
            ]
        )
        fill = fill_by_status.get(record.match_status)
        if fill:
            for cell in bank_sheet[bank_sheet.max_row]:
                cell.fill = fill
    workbook.save(output_path)


MATCH_TEXT_STOP_WORDS = {
    "aed",
    "are",
    "abu",
    "dhabi",
    "dubai",
    "sharjah",
    "alain",
    "al",
    "transaction",
    "transactions",
    "transfer",
    "transfers",
    "trf",
    "mbtrf",
    "mobn",
    "from",
    "to",
    "out",
    "within",
    "charges",
    "incl",
    "vat",
    "debit",
    "credit",
    "payment",
    "paymnt",
    "received",
    "top",
    "up",
    "reference",
    "number",
    "card",
    "creditcard",
    "purchase",
    "purchaseepp",
    "thank",
    "you",
    "thankyou",
    "current",
    "account",
    "cash",
    "deposit",
    "cheque",
    "salary",
    "profit",
    "reversal",
    "fee",
    "fees",
    "wallet",
    "branch",
    "site",
}
RECEIPT_NOT_REQUIRED_PATTERNS = (
    r"\btransfer\b",
    r"\bmbtrf\b",
    r"\bmobn\b",
    r"\bpayment received\b",
    r"\btop up\b",
    r"\bcashback\b",
    r"\bcheque deposit\b",
    r"\bcash deposit\b",
    r"\bsalary transfer\b",
    r"\bprofit credit\b",
    r"\breversal\b",
    r"\bcredit card paym",
    r"\batm wdl\b",
    r"\bfee\b",
    r"\bcharges?\b",
)


def amount_to_float(value: str) -> Optional[float]:
    cleaned = normalize_whitespace(str(value or "")).replace(",", "")
    if not cleaned or cleaned == "Unknown":
        return None
    cleaned = cleaned.replace("AED", "").replace("aed", "").strip()
    try:
        return abs(float(cleaned))
    except ValueError:
        return None


def parse_sheet_transaction_direction(row: dict[str, Any]) -> str:
    def numeric(value: Any) -> Optional[float]:
        if value in (None, ""):
            return None
        cleaned = str(value).replace(",", "").strip()
        cleaned = cleaned.replace("AED", "").replace("aed", "").strip()
        cleaned = cleaned.replace("CR", "").replace("cr", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return None

    for key in ("debit", "withdrawal", "dr"):
        value = numeric(row.get(key))
        if value is not None and abs(value) > 0:
            return "debit"
    for key in ("credit", "deposit", "cr"):
        value = numeric(row.get(key))
        if value is not None and abs(value) > 0:
            return "credit"
    amount_value = numeric(row.get("amount"))
    if amount_value is not None:
        if amount_value < 0:
            return "debit"
        if amount_value > 0:
            return "credit"
    return "unknown"


def parse_record_date(value: str) -> Optional[datetime]:
    cleaned = normalize_whitespace(str(value or ""))
    if not cleaned or cleaned == "Unknown":
        return None
    try:
        return datetime.strptime(cleaned, "%Y-%m-%d")
    except ValueError:
        return None


def normalize_match_tokens(value: str) -> List[str]:
    normalized = normalize_ocr_text(normalize_whitespace(value)).lower()
    normalized = re.sub(r"https?://\S+", " ", normalized)
    tokens = re.findall(r"[a-z]{2,}", normalized)
    return [token for token in tokens if token not in MATCH_TEXT_STOP_WORDS]


def vendor_similarity(left: str, right: str) -> float:
    left_tokens = normalize_match_tokens(left)
    right_tokens = normalize_match_tokens(right)
    if not left_tokens or not right_tokens:
        return 0.0
    left_joined = " ".join(left_tokens)
    right_joined = " ".join(right_tokens)
    sequence_score = SequenceMatcher(None, left_joined, right_joined).ratio()
    left_set = set(left_tokens)
    right_set = set(right_tokens)
    overlap_score = len(left_set & right_set) / max(len(left_set), len(right_set))
    containment_score = 1.0 if (left_joined in right_joined or right_joined in left_joined) else 0.0
    return max(sequence_score, overlap_score, containment_score)


def bank_transaction_requires_receipt(record: ProcessedRecord) -> Tuple[bool, str]:
    if record.transaction_direction == "credit":
        return False, "credit transaction"
    haystack = normalize_ocr_text(" ".join([record.company_name, record.raw_text])).lower()
    for pattern in RECEIPT_NOT_REQUIRED_PATTERNS:
        if re.search(pattern, haystack, flags=re.IGNORECASE):
            return False, f"excluded by pattern: {pattern}"
    return True, ""


def score_record_match(bank_record: ProcessedRecord, candidate_record: ProcessedRecord) -> Tuple[int, str]:
    bank_amount = amount_to_float(bank_record.amount)
    candidate_amount = amount_to_float(candidate_record.amount)
    if bank_amount is None or candidate_amount is None:
        return 0, "amount missing"

    amount_delta = abs(bank_amount - candidate_amount)
    if amount_delta <= 0.05:
        amount_score = 55
        amount_reason = "amount exact"
    elif amount_delta <= 1.0:
        amount_score = 35
        amount_reason = "amount close"
    else:
        return 0, "amount mismatch"

    similarity = vendor_similarity(bank_record.company_name, candidate_record.company_name)
    if similarity < 0.35:
        return 0, f"vendor weak {similarity:.2f}"
    vendor_score = int(round(similarity * 30))

    date_score = 0
    date_reason = "date unknown"
    bank_date = parse_record_date(bank_record.date)
    candidate_date = parse_record_date(candidate_record.date)
    if bank_date and candidate_date:
        day_delta = abs((bank_date - candidate_date).days)
        if day_delta == 0:
            date_score = 15
            date_reason = "same date"
        elif day_delta <= 3:
            date_score = 10
            date_reason = f"{day_delta}d apart"
        elif day_delta <= 7:
            date_score = 5
            date_reason = f"{day_delta}d apart"
        else:
            date_reason = f"{day_delta}d apart"

    doc_bonus = 5 if candidate_record.doc_type in {"Invoice", "Receipt", "Purchase"} else 0
    total_score = amount_score + vendor_score + date_score + doc_bonus
    if amount_delta <= 0.05:
        minimum_score = 70
    else:
        minimum_score = 78
    if total_score < minimum_score:
        return 0, f"below threshold {total_score}"
    return total_score, f"{amount_reason}; vendor {similarity:.2f}; {date_reason}"


def reconcile_bank_transactions(records: Sequence[ProcessedRecord]) -> None:
    bank_indexes: List[int] = []
    candidate_indexes: List[int] = []

    for index, record in enumerate(records):
        record.match_score = 0
        record.matched_record_source_file = ""
        record.matched_record_output_file = ""
        record.matched_record_source_type = ""
        record.matched_record_source_timestamp = ""
        record.matched_record_date = ""
        record.matched_record_number = ""
        record.matched_record_company_name = ""
        record.matched_record_amount = ""
        record.matched_record_transaction_direction = ""
        record.match_basis = ""
        if record.source_type == "sheet" and record.doc_type == "BankTransaction":
            requires_receipt, reason = bank_transaction_requires_receipt(record)
            if requires_receipt:
                record.match_status = "missing_receipt"
                bank_indexes.append(index)
            else:
                record.match_status = "not_applicable"
                record.match_basis = reason
        else:
            record.match_status = "candidate"
            if record.source_type in {"pdf", "video"} and amount_to_float(record.amount) is not None:
                candidate_indexes.append(index)

    scored_pairs: List[Tuple[int, int, int, str]] = []
    for bank_index in bank_indexes:
        for candidate_index in candidate_indexes:
            score, basis = score_record_match(records[bank_index], records[candidate_index])
            if score > 0:
                scored_pairs.append((score, bank_index, candidate_index, basis))

    matched_banks: set[int] = set()
    matched_candidates: set[int] = set()
    for score, bank_index, candidate_index, basis in sorted(scored_pairs, reverse=True):
        if bank_index in matched_banks or candidate_index in matched_candidates:
            continue
        bank_record = records[bank_index]
        candidate_record = records[candidate_index]
        bank_record.match_status = "matched"
        bank_record.match_score = score
        bank_record.matched_record_source_file = candidate_record.source_file
        bank_record.matched_record_output_file = candidate_record.output_file
        bank_record.matched_record_source_type = candidate_record.source_type
        bank_record.matched_record_source_timestamp = candidate_record.source_timestamp
        bank_record.matched_record_date = candidate_record.date
        bank_record.matched_record_number = candidate_record.number
        bank_record.matched_record_company_name = candidate_record.company_name
        bank_record.matched_record_amount = candidate_record.amount
        bank_record.matched_record_transaction_direction = candidate_record.transaction_direction
        bank_record.match_basis = basis

        candidate_record.match_status = "linked_to_bank"
        candidate_record.match_score = score
        candidate_record.matched_record_source_file = bank_record.source_file
        candidate_record.matched_record_output_file = bank_record.output_file
        candidate_record.matched_record_source_type = bank_record.source_type
        candidate_record.matched_record_source_timestamp = bank_record.source_timestamp
        candidate_record.matched_record_date = bank_record.date
        candidate_record.matched_record_number = bank_record.number
        candidate_record.matched_record_company_name = bank_record.company_name
        candidate_record.matched_record_amount = bank_record.amount
        candidate_record.matched_record_transaction_direction = bank_record.transaction_direction
        candidate_record.match_basis = basis

        matched_banks.add(bank_index)
        matched_candidates.add(candidate_index)


def image_to_pdf(image: np.ndarray, output_path: Path) -> None:
    height, width = image.shape[:2]
    success, encoded = cv2.imencode(".png", image)
    if not success:
        raise RuntimeError("Failed to encode image for PDF output.")
    pdf = fitz.open()
    page = pdf.new_page(width=width, height=height)
    page.insert_image(page.rect, stream=encoded.tobytes())
    pdf.save(output_path)
    pdf.close()


def image_to_png(image: np.ndarray, output_path: Path) -> None:
    success, encoded = cv2.imencode(".png", image)
    if not success:
        raise RuntimeError("Failed to encode image for PNG output.")
    output_path.write_bytes(encoded.tobytes())


def ensure_bgr(image: np.ndarray) -> np.ndarray:
    if len(image.shape) == 2:
        return cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
    return image


def iter_pdf_files(source_dir: Path) -> Iterable[Path]:
    yield from sorted(path for path in source_dir.glob("*.pdf") if path.is_file())


def iter_source_files(source_dir: Path) -> Iterable[Path]:
    supported = {".pdf", ".mp4", ".mov", ".m4v", ".avi", ".xlsx", ".xlsm", ".csv"}
    yield from sorted(
        path for path in source_dir.iterdir()
        if path.is_file() and path.suffix.lower() in supported
    )


def detect_source_file_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return "pdf"
    if suffix in {".mp4", ".mov", ".m4v", ".avi"}:
        return "video"
    if suffix in {".xlsx", ".xlsm", ".csv"}:
        return "sheet"
    return "unknown"


def format_seconds_timestamp(seconds: float) -> str:
    total_millis = max(0, int(round(seconds * 1000)))
    hours, remainder = divmod(total_millis, 3_600_000)
    minutes, remainder = divmod(remainder, 60_000)
    secs, millis = divmod(remainder, 1000)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}.{millis:03d}"


def infer_source_origin(source_path: Path, source_type: str) -> str:
    name = source_path.name.lower()
    if source_type == "video":
        if "tabby" in name:
            return "tabby_video"
        if "amazon" in name:
            return "amazon_video"
        if "whatsapp" in name:
            return "whatsapp_video"
        return "video_upload"
    return "pdf_upload"


VIDEO_DATE_LINE_RE = re.compile(
    r"(?i)\b(\d{1,2})\s*"
    r"(jan|january|feb|february|mar|march|apr|april|may|jun|june|jul|july|aug|august|"
    r"sep|sept|september|oct|october|nov|november|dec|december)"
    r"(?:\s+(\d{4}))?\b"
)
VIDEO_FILTER_TERMS = {
    "allproducts",
    "date",
    "history",
    "products",
}
VIDEO_META_TERMS = {
    "petrolstations",
    "groceries",
    "dining",
    "governmentpayments",
    "financialservices",
    "other",
    "clothingshoes",
    "clothingandshoes",
    "payment",
    "tabbycard",
    "cardrepayment",
}


def video_content_bounds(image: np.ndarray) -> BoundingBox:
    height, width = image.shape[:2]
    x1 = int(width * 0.05)
    y1 = int(height * 0.18)
    x2 = int(width * 0.97)
    y2 = int(height * 0.93)
    return x1, y1, max(1, x2 - x1), max(1, y2 - y1)


def frame_signature(image: np.ndarray) -> str:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    tiny = cv2.resize(gray, (16, 16), interpolation=cv2.INTER_AREA)
    mean_value = float(np.mean(tiny))
    bits = ["1" if value >= mean_value else "0" for value in tiny.flatten()]
    return "".join(bits)


def signature_distance(left: str, right: str) -> int:
    return sum(1 for a, b in zip(left, right) if a != b)


def detect_video_record_regions(image: np.ndarray) -> List[BoundingBox]:
    height, width = image.shape[:2]
    x, y, w, h = video_content_bounds(image)
    if w <= 1 or h <= 1:
        return [(0, 0, width, height)]

    region = image[y:y + h, x:x + w]
    hsv = cv2.cvtColor(region, cv2.COLOR_BGR2HSV)
    gray = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
    non_white = ((gray < 245) | ((hsv[:, :, 1] > 18) & (hsv[:, :, 2] < 250))).astype(np.uint8) * 255
    non_white = cv2.morphologyEx(
        non_white,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(cv2.MORPH_RECT, (21, 3)),
    )
    non_white = cv2.morphologyEx(
        non_white,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3)),
    )
    projection = np.sum(non_white > 0, axis=1)
    threshold = max(8, int(w * 0.02))

    boxes: List[BoundingBox] = []
    start: Optional[int] = None
    quiet_rows = 0
    max_band_height = int(height * 0.16)
    min_band_height = int(height * 0.022)

    for index, value in enumerate(projection):
        if value > threshold:
            if start is None:
                start = index
            quiet_rows = 0
        elif start is not None:
            quiet_rows += 1
            if quiet_rows > 6:
                end = index - quiet_rows + 1
                band_height = end - start
                if min_band_height <= band_height <= max_band_height:
                    boxes.append((x, y + start, w, band_height))
                start = None
                quiet_rows = 0

    if start is not None:
        end = len(projection)
        band_height = end - start
        if min_band_height <= band_height <= max_band_height:
            boxes.append((x, y + start, w, band_height))

    merged = merge_boxes(boxes, gap=8)
    filtered = [
        box for box in merged
        if min_band_height <= box[3] <= max_band_height
    ]
    if filtered:
        return filtered[:14]

    estimated_height = max(52, int(height * 0.08))
    estimated_gap = max(8, int(height * 0.012))
    fallback: List[BoundingBox] = []
    cursor = y
    while cursor + estimated_height <= y + h and len(fallback) < 8:
        fallback.append((x, cursor, w, estimated_height))
        cursor += estimated_height + estimated_gap
    return fallback or [(x, y, w, min(h, estimated_height))]


def parse_video_date_header(text: str) -> str:
    normalized = normalize_ocr_text(text)
    for line in text_lines(normalized):
        match = VIDEO_DATE_LINE_RE.search(line)
        if not match:
            continue
        day = match.group(1)
        month = match.group(2)
        year = match.group(3) or str(datetime.now().year)
        for fmt in ("%d %B %Y", "%d %b %Y"):
            try:
                return datetime.strptime(f"{day} {month} {year}", fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return "Unknown"


def is_video_filter_text(text: str) -> bool:
    joined = canonical_letters(text)
    if not joined:
        return False
    return any(term in joined for term in VIDEO_FILTER_TERMS)


def clean_video_company_name(text: str) -> str:
    candidates: List[str] = []
    for line in text_lines(normalize_ocr_text(text)):
        canonical = canonical_letters(line)
        if not canonical or canonical in VIDEO_FILTER_TERMS:
            continue
        if canonical in VIDEO_META_TERMS:
            continue
        if VIDEO_DATE_LINE_RE.search(line):
            continue
        if re.search(r"(?i)\bpayment\s+\d+\s+of\s+\d+\b", line):
            continue
        if re.search(r"(?i)\b(?:history|all\s+products|date)\b", line):
            continue
        cleaned = re.sub(r"\s{2,}", " ", line).strip(" -:")
        if len(cleaned) < 3:
            continue
        candidates.append(cleaned)
    return candidates[0][:120] if candidates else "Unknown"


def parse_video_amount(text: str) -> str:
    normalized = normalize_ocr_text(text).replace(",", "")
    matches = re.findall(r"(?i)[+\-]?\s*[ÐAED]*\s*(\d{1,6}\.\d{2})", normalized)
    if not matches:
        return "Unknown"
    try:
        return f"{float(matches[0]):.2f}"
    except ValueError:
        return "Unknown"


def parse_video_reference(text: str) -> str:
    normalized = normalize_ocr_text(text)
    payment_match = re.search(r"(?i)\bpayment\s+(\d+\s+of\s+\d+)\b", normalized)
    if payment_match:
        return payment_match.group(1).replace(" ", "")
    card_match = re.search(r"(?i)\b(?:tabby\s+card|card\s+repayment)\b.*?([*•\.]{2,}\s*\d{4})", normalized)
    if card_match:
        digits = re.sub(r"\D", "", card_match.group(1))
        if digits:
            return f"card-{digits}"
    return "Unknown"


def detect_video_date_anchors(
    image: np.ndarray,
    ocr_backend: str,
    ocr_engines: Sequence[Any],
) -> List[Tuple[int, str]]:
    x, y, w, h = video_content_bounds(image)
    left_width = max(1, int(w * 0.56))
    stripe_height = max(70, int(image.shape[0] * 0.08))
    step = max(36, int(image.shape[0] * 0.055))
    anchors: List[Tuple[int, str]] = []

    cursor = y
    while cursor + stripe_height <= y + h:
        stripe = image[cursor:cursor + stripe_height, x:x + left_width]
        text, _ = extract_text_by_backend(stripe, ocr_backend, ocr_engines)
        parsed = parse_video_date_header(text)
        if parsed != "Unknown":
            center_y = cursor + stripe_height // 2
            if not anchors or anchors[-1][1] != parsed or abs(anchors[-1][0] - center_y) > step:
                anchors.append((center_y, parsed))
        cursor += step
    return anchors


def extract_video_row_fields(
    full_text: str,
    left_text: str,
    right_text: str,
    current_date: str,
) -> DocumentFields:
    combined = normalize_ocr_text("\n".join(part for part in [full_text, left_text, right_text] if part.strip()))
    company_name = clean_video_company_name(left_text)
    if company_name == "Unknown":
        company_name = clean_video_company_name(combined)
    amount = parse_video_amount(right_text)
    if amount == "Unknown":
        amount = parse_video_amount(combined)
    number = parse_video_reference(combined)
    return DocumentFields(
        doc_type="Purchase" if company_name != "Unknown" or amount != "Unknown" else "Unknown",
        date=current_date,
        number=number,
        company_name=company_name,
        amount=amount,
    )


def process_video(
    video_path: Path,
    output_dir: Path,
    project_name: str,
    ocr_engines: Sequence[Any],
    ocr_backend: str,
    sample_seconds: int = 2,
    max_frames: int = 120,
    debug_image_dir: Optional[Path] = None,
    trocr_engine: Any = None,
    naming_pattern: str = DEFAULT_NAMING_PATTERN,
    log_message: Optional[Any] = None,
    item_complete: Optional[Any] = None,
    should_cancel: Optional[Any] = None,
) -> Tuple[List[Path], List[ProcessedRecord]]:
    def emit(message: str) -> None:
        if log_message:
            log_message(message)
        else:
            print(message)

    capture = cv2.VideoCapture(str(video_path))
    if not capture.isOpened():
        raise RuntimeError(f"Unable to open video: {video_path}")

    fps = capture.get(cv2.CAP_PROP_FPS) or 30.0
    frame_count = int(capture.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
    duration = frame_count / fps if frame_count > 0 and fps > 0 else 0
    source_hash = compute_file_sha256(video_path)
    source_origin = infer_source_origin(video_path, "video")

    step_frames = max(1, int(round(max(sample_seconds, 1) * fps)))
    if frame_count > 0 and max_frames > 0:
        step_frames = max(step_frames, max(1, frame_count // max_frames))

    generated_files: List[Path] = []
    records: List[ProcessedRecord] = []
    previous_signature = ""
    seen_record_keys: set[Tuple[str, str, str, str, str]] = set()
    current_date = "Unknown"

    frame_index = 0
    sampled_index = 0
    while True:
        if should_cancel and should_cancel():
            emit("Run cancelled before processing next video frame.")
            break
        capture.set(cv2.CAP_PROP_POS_FRAMES, frame_index)
        ok, frame = capture.read()
        if not ok or frame is None:
            break

        sampled_index += 1
        emit(f"Sampled frame {sampled_index} at {format_seconds_timestamp(frame_index / fps if fps > 0 else 0.0)}")
        content_signature = frame_signature(crop_with_padding(frame, video_content_bounds(frame), padding=0))
        if previous_signature and signature_distance(previous_signature, content_signature) < 20:
            frame_index += step_frames
            continue
        previous_signature = content_signature

        timestamp_seconds = frame_index / fps if fps > 0 else 0.0
        timestamp_label = format_seconds_timestamp(timestamp_seconds)
        regions = detect_video_record_regions(frame)
        date_anchors = detect_video_date_anchors(frame, ocr_backend, ocr_engines)

        for region_index, box in enumerate(regions, start=1):
            crop = crop_with_padding(frame, box, padding=10)
            text, enhanced_image = extract_text_by_backend(crop, ocr_backend, ocr_engines)
            normalized_text = normalize_ocr_text(text)
            if len(normalized_text.strip()) < 8:
                continue

            if is_video_filter_text(normalized_text):
                continue

            header_date = parse_video_date_header(normalized_text)
            if header_date != "Unknown" and parse_video_amount(normalized_text) == "Unknown":
                current_date = header_date
                emit(f"Detected video date header: {current_date}")
                continue
            inherited_date = current_date
            region_center = box[1] + (box[3] // 2)
            for anchor_y, anchor_date in date_anchors:
                if anchor_y <= region_center:
                    inherited_date = anchor_date

            left_crop = crop_relative(crop, 0.14, 0.0, 0.69, 1.0)
            right_crop = crop_relative(crop, 0.68, 0.0, 0.99, 1.0)
            left_text, _ = extract_text_by_backend(left_crop, ocr_backend, ocr_engines)
            right_text, _ = extract_text_by_backend(right_crop, ocr_backend, ocr_engines)
            fields = extract_video_row_fields(
                full_text=normalized_text,
                left_text=left_text,
                right_text=right_text,
                current_date=header_date if header_date != "Unknown" else inherited_date,
            )
            if header_date != "Unknown":
                current_date = header_date

            if fields.company_name == "Unknown" or fields.amount == "Unknown":
                continue
            record_key = (
                fields.doc_type,
                fields.date,
                fields.number,
                fields.company_name,
                fields.amount,
            )
            if record_key in seen_record_keys:
                continue
            seen_record_keys.add(record_key)

            filename = build_output_name(fields, project_name, naming_pattern)
            confidence_score, confidence_label = compute_confidence(fields)
            output_path = output_dir / filename

            duplicate_counter = 1
            while output_path.exists():
                output_path = output_dir / (
                    filename[:-4] + f"_t{timestamp_label.replace(':', '').replace('.', '')}_{region_index}_{duplicate_counter}.pdf"
                )
                duplicate_counter += 1

            image_to_pdf(crop, output_path)
            generated_files.append(output_path)
            records.append(
                ProcessedRecord(
                    source_file=video_path.name,
                    source_path=str(video_path.resolve()),
                    source_hash=source_hash,
                    source_type="video",
                    source_origin=source_origin,
                    source_timestamp=timestamp_label,
                    output_file=output_path.name,
                    doc_type=fields.doc_type,
                    date=fields.date,
                    number=fields.number,
                    company_name=fields.company_name,
                    amount=fields.amount,
                    currency="AED" if fields.amount != "Unknown" else "Unknown",
                    transaction_direction="debit",
                    project_name=project_name,
                    confidence_score=confidence_score,
                    confidence_label=confidence_label,
                    raw_text=normalize_ocr_text(
                        "\n".join([normalized_text, normalize_ocr_text(left_text), normalize_ocr_text(right_text)])
                    ),
                )
            )
            if item_complete:
                item_complete(video_path, [output_path], [records[-1]])
            emit(
                f"Extracted video record at {timestamp_label}: "
                f"{fields.company_name} {fields.amount}"
            )

            if debug_image_dir is not None:
                debug_image_dir.mkdir(parents=True, exist_ok=True)
                debug_base = f"{video_path.stem}_t{timestamp_label.replace(':', '-').replace('.', '-')}_r{region_index}"
                image_to_png(crop, debug_image_dir / f"{debug_base}_original.png")
                image_to_png(ensure_bgr(enhanced_image), debug_image_dir / f"{debug_base}_enhanced.png")

        frame_index += step_frames
        if max_frames > 0 and sampled_index >= max_frames:
            break

    capture.release()
    return generated_files, records


def parse_sheet_date(value: Any) -> str:
    if value is None:
        return "Unknown"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = normalize_ocr_text(str(value))
    return parse_date(text)


def parse_sheet_amount(row: dict[str, Any]) -> str:
    def numeric(value: Any) -> Optional[float]:
        if value in (None, ""):
            return None
        cleaned = str(value).replace(",", "").strip()
        cleaned = cleaned.replace("AED", "").replace("aed", "").strip()
        try:
            return abs(float(cleaned))
        except ValueError:
            return None

    for key in ("amount", "transaction_amount"):
        value = numeric(row.get(key))
        if value is not None and value > 0:
            return f"{value:.2f}"
    for key in ("debit", "withdrawal", "dr"):
        value = numeric(row.get(key))
        if value is not None and value > 0:
            return f"{value:.2f}"
    for key in ("credit", "deposit", "cr"):
        value = numeric(row.get(key))
        if value is not None and value > 0:
            return f"{value:.2f}"
    return "Unknown"


def normalize_sheet_headers(values: Sequence[Any]) -> dict[int, str]:
    mappings = {
        "date": {"date", "transaction date", "value date", "posting date"},
        "description": {
            "description",
            "details",
            "merchant",
            "narration",
            "transaction",
            "transaction no.",
            "transaction no",
            "remarks",
            "notes",
            "reference",
            "type",
            "transaction type",
        },
        "reference": {
            "reference",
            "ref",
            "transaction id",
            "transaction no",
            "transaction no.",
            "order id",
            "reference no",
            "reference number",
            "reference no.",
            "ref. number",
            "original ref. number",
            "sl.no",
            "sl no",
        },
        "amount": {"amount", "transaction amount", "local currency"},
        "debit": {"debit", "debit amount", "withdrawal", "dr", "debit(-)"},
        "credit": {"credit", "credit amount", "deposit", "cr", "credit(-)"},
        "currency": {"currency", "curr"},
    }
    header_map: dict[int, str] = {}
    for index, value in enumerate(values):
        text = normalize_whitespace(str(value or "")).lower()
        for normalized, aliases in mappings.items():
            if text in aliases:
                header_map[index] = normalized
                break
    return header_map


def build_sheet_row_data(values: Sequence[Any], header_map: dict[int, str]) -> dict[str, Any]:
    row_data: dict[str, Any] = {}
    for column_index, normalized in header_map.items():
        value = values[column_index] if column_index < len(values) else ""
        if normalized not in row_data:
            row_data[normalized] = value
            continue
        existing = row_data.get(normalized)
        if existing in (None, "") and value not in (None, ""):
            row_data[normalized] = value
    return row_data


def create_summary_pdf(lines: Sequence[str], output_path: Path, width: int = 1240) -> None:
    font = ImageFont.load_default()
    padding = 40
    line_height = 26
    usable_lines = [line[:140] for line in lines if line][:30]
    height = max(400, padding * 2 + line_height * (len(usable_lines) + 2))
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    draw.text((padding, padding), "ULTRA FORCE Extracted Record", fill="black", font=font)
    cursor_y = padding + 46
    for line in usable_lines:
        draw.text((padding, cursor_y), line, fill="black", font=font)
        cursor_y += line_height
    image_to_pdf(cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR), output_path)


def process_sheet(
    sheet_path: Path,
    output_dir: Path,
    project_name: str,
    naming_pattern: str = DEFAULT_NAMING_PATTERN,
    log_message: Optional[Any] = None,
    item_complete: Optional[Any] = None,
) -> Tuple[List[Path], List[ProcessedRecord]]:
    def emit(message: str) -> None:
        if log_message:
            log_message(message)
        else:
            print(message)

    source_hash = compute_file_sha256(sheet_path)
    source_origin = "csv_import" if sheet_path.suffix.lower() == ".csv" else "excel_import"
    sheets: List[Tuple[str, List[List[Any]]]] = []
    if sheet_path.suffix.lower() == ".csv":
        with sheet_path.open("r", encoding="utf-8-sig", newline="") as handle:
            sheets.append((sheet_path.stem, list(csv.reader(handle))))
    else:
        workbook = load_workbook(sheet_path, data_only=True, read_only=True)
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            sheets.append((worksheet.title, rows))

    generated_files: List[Path] = []
    records: List[ProcessedRecord] = []
    recognized_sheet = False
    for sheet_name, rows in sheets:
        header_index = None
        header_map: dict[int, str] = {}
        for index, row in enumerate(rows[:15]):
            current_map = normalize_sheet_headers(row)
            if "date" in current_map.values() and ("description" in current_map.values() or "amount" in current_map.values()):
                header_index = index
                header_map = current_map
                recognized_sheet = True
                break
        if header_index is None:
            emit(f"Skipped sheet tab: {sheet_path.name} [{sheet_name}] -> no recognizable header row")
            continue
        for row_index, values in enumerate(rows[header_index + 1 :], start=1):
            row_data = build_sheet_row_data(values, header_map)
            raw_text = " | ".join(normalize_whitespace(str(value or "")) for value in values if value not in (None, ""))
            if not raw_text.strip():
                continue
            fields = DocumentFields(
                doc_type="BankTransaction",
                date=parse_sheet_date(row_data.get("date")),
                number=normalize_whitespace(str(row_data.get("reference") or "")) or "Unknown",
                company_name=normalize_whitespace(str(row_data.get("description") or ""))[:120] or "Unknown",
                amount=parse_sheet_amount(row_data),
            )
            if fields.date == "Unknown" and fields.company_name == "Unknown" and fields.amount == "Unknown":
                continue
            confidence_score, confidence_label = compute_confidence(fields)
            filename = build_output_name(fields, project_name, naming_pattern)
            output_path = output_dir / filename
            duplicate_counter = 1
            sheet_suffix = sanitize_filename_part(sheet_name)
            while output_path.exists():
                output_path = output_dir / (
                    filename[:-4] + f"_{sheet_suffix}_row{row_index}_{duplicate_counter}.pdf"
                )
                duplicate_counter += 1

            create_summary_pdf(
                [
                    f"Source: {sheet_path.name}",
                    f"Sheet: {sheet_name}",
                    f"Row: {row_index}",
                    f"Type: {fields.doc_type}",
                    f"Date: {fields.date}",
                    f"Reference: {fields.number}",
                    f"Company: {fields.company_name}",
                    f"Amount: {fields.amount}",
                    f"Raw: {raw_text}",
                ],
                output_path,
            )
            generated_files.append(output_path)
            record = ProcessedRecord(
                source_file=sheet_path.name,
                source_path=str(sheet_path.resolve()),
                source_hash=source_hash,
                source_type="sheet",
                source_origin=source_origin,
                source_timestamp=f"{sheet_name}:row:{row_index}",
                output_file=output_path.name,
                doc_type=fields.doc_type,
                date=fields.date,
                number=fields.number,
                company_name=fields.company_name,
                amount=fields.amount,
                currency=normalize_whitespace(str(row_data.get("currency") or "")) or ("AED" if fields.amount != "Unknown" else "Unknown"),
                transaction_direction=parse_sheet_transaction_direction(row_data),
                project_name=project_name,
                confidence_score=confidence_score,
                confidence_label=confidence_label,
                raw_text=f"Sheet: {sheet_name} | {raw_text}",
            )
            records.append(record)
            if item_complete:
                item_complete(sheet_path, [output_path], [record])

        emit(f"Processed sheet tab: {sheet_path.name} [{sheet_name}]")

    if not recognized_sheet:
        emit(f"Skipped sheet: {sheet_path.name} -> no recognizable header row")
        return [], []

    emit(f"Processed sheet workbook: {sheet_path.name}")
    return generated_files, records


def process_pdf(
    pdf_path: Path,
    output_dir: Path,
    project_name: str,
    ocr_engines: Sequence[Any],
    ocr_backend: str,
    dpi: int = 300,
    save_text: bool = False,
    single_item_per_page: bool = False,
    export_image_mode: str = "original",
    debug_image_dir: Optional[Path] = None,
    trocr_engine: Any = None,
    naming_pattern: str = DEFAULT_NAMING_PATTERN,
) -> Tuple[List[Path], List[ProcessedRecord]]:
    output_paths: List[Path] = []
    records: List[ProcessedRecord] = []
    pages = render_pdf_to_images(pdf_path, dpi=dpi)
    source_hash = compute_file_sha256(pdf_path)

    for page_index, page_image in enumerate(pages, start=1):
        if single_item_per_page:
            boxes = [(0, 0, page_image.shape[1], page_image.shape[0])]
        else:
            boxes = detect_document_regions(page_image)

        for item_index, box in enumerate(boxes, start=1):
            cropped = crop_with_padding(page_image, box)
            text, enhanced_image = extract_text_by_backend(cropped, ocr_backend, ocr_engines)
            fields = extract_fields(text)
            fields = refine_fields_with_region_ocr(
                cropped,
                fields,
                ocr_backend=ocr_backend,
                ocr_engines=ocr_engines,
                trocr_engine=trocr_engine,
            )

            filename = build_output_name(fields, project_name, naming_pattern)
            confidence_score, confidence_label = compute_confidence(fields)
            output_path = output_dir / filename

            duplicate_counter = 1
            while output_path.exists():
                output_path = output_dir / (
                    filename[:-4]
                    + f"_p{page_index}_i{item_index}_{duplicate_counter}.pdf"
                )
                duplicate_counter += 1

            exports: List[Tuple[np.ndarray, Path]] = []
            if export_image_mode in {"original", "both"}:
                exports.append((cropped, output_path))
            if export_image_mode in {"enhanced", "both"}:
                enhanced_output = output_path
                if export_image_mode == "both":
                    enhanced_output = output_dir / (output_path.stem + "_enhanced.pdf")
                exports.append((ensure_bgr(enhanced_image), enhanced_output))

            for export_image, export_path in exports:
                image_to_pdf(export_image, export_path)
                output_paths.append(export_path)
                records.append(
                    ProcessedRecord(
                        source_file=pdf_path.name,
                        source_path=str(pdf_path.resolve()),
                        source_hash=source_hash,
                        source_type="pdf",
                        source_origin="pdf_upload",
                        source_timestamp="",
                        output_file=export_path.name,
                        doc_type=fields.doc_type,
                        date=fields.date,
                        number=fields.number,
                        company_name=fields.company_name,
                        amount=fields.amount,
                        currency="AED" if fields.amount != "Unknown" else "Unknown",
                        transaction_direction="debit",
                        project_name=project_name,
                        confidence_score=confidence_score,
                        confidence_label=confidence_label,
                        raw_text=text,
                    )
                )
            if save_text:
                text_path = output_path.with_suffix(".txt")
                text_path.write_text(text, encoding="utf-8")

            if debug_image_dir is not None:
                debug_image_dir.mkdir(parents=True, exist_ok=True)
                debug_base = (
                    f"{pdf_path.stem}_p{page_index}_i{item_index}"
                )
                image_to_png(cropped, debug_image_dir / f"{debug_base}_original.png")
                image_to_png(
                    ensure_bgr(enhanced_image),
                    debug_image_dir / f"{debug_base}_enhanced.png",
                )

    return output_paths, records


def process_folder(
    source_dir: Path,
    output_dir: Path,
    project_name: str,
    dpi: int = 300,
    language: str = "en",
    use_angle_cls: bool = True,
    save_text: bool = False,
    ocr_profile: str = "mixed",
    single_item_per_page: bool = False,
    ocr_backend: str = "ai",
    export_image_mode: str = "original",
    debug_image_dir: Optional[Path] = None,
    handwriting_backend: str = "none",
    trocr_model: str = "microsoft/trocr-base-handwritten",
    video_sample_seconds: int = 2,
    video_max_frames: int = 120,
    naming_pattern: str = DEFAULT_NAMING_PATTERN,
    log_message: Optional[Any] = None,
    item_complete: Optional[Any] = None,
    should_cancel: Optional[Any] = None,
    should_skip: Optional[Any] = None,
) -> Tuple[List[Path], List[ProcessedRecord]]:
    def emit(message: str) -> None:
        if log_message:
            log_message(message)
        else:
            print(message)

    output_dir.mkdir(parents=True, exist_ok=True)
    ocr_engines: Sequence[Any] = []
    trocr_engine: Any = None
    if ocr_backend == "ai":
        ocr_engines = make_ocr_engines(
            language=language,
            use_angle_cls=use_angle_cls,
            ocr_profile=ocr_profile,
        )
    if handwriting_backend == "trocr":
        trocr_engine = make_trocr_engine(trocr_model)
    generated_files: List[Path] = []
    records: List[ProcessedRecord] = []

    for source_path in iter_source_files(source_dir):
        if should_cancel and should_cancel():
            emit("Run cancelled before processing next file.")
            break
        if should_skip and should_skip(source_path):
            emit(f"Skipped duplicate: {source_path.name}")
            continue
        try:
            source_type = detect_source_file_type(source_path)
            if source_type == "pdf":
                file_outputs, file_records = process_pdf(
                    pdf_path=source_path,
                    output_dir=output_dir,
                    project_name=project_name,
                    ocr_engines=ocr_engines,
                    ocr_backend=ocr_backend,
                    dpi=dpi,
                    save_text=save_text,
                    single_item_per_page=single_item_per_page,
                    export_image_mode=export_image_mode,
                    debug_image_dir=debug_image_dir,
                    trocr_engine=trocr_engine,
                    naming_pattern=naming_pattern,
                )
            elif source_type == "video":
                file_outputs, file_records = process_video(
                    video_path=source_path,
                    output_dir=output_dir,
                    project_name=project_name,
                    ocr_engines=ocr_engines,
                    ocr_backend=ocr_backend,
                    sample_seconds=video_sample_seconds,
                    max_frames=video_max_frames,
                    debug_image_dir=debug_image_dir,
                    trocr_engine=trocr_engine,
                    naming_pattern=naming_pattern,
                    log_message=emit,
                    item_complete=item_complete,
                    should_cancel=should_cancel,
                )
            elif source_type == "sheet":
                file_outputs, file_records = process_sheet(
                    sheet_path=source_path,
                    output_dir=output_dir,
                    project_name=project_name,
                    naming_pattern=naming_pattern,
                    log_message=emit,
                    item_complete=item_complete,
                )
            else:
                emit(f"Skipped unsupported file: {source_path.name}")
                continue
            generated_files.extend(file_outputs)
            records.extend(file_records)
            if item_complete and source_type == "pdf":
                item_complete(source_path, file_outputs, file_records)
            if source_type != "video" and source_type != "sheet":
                emit(f"Processed: {source_path.name}")
        except Exception as exc:
            emit(f"Failed: {source_path.name} -> {exc}")
    return generated_files, records


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Split scanned invoice PDFs, OCR each item, extract fields, and save renamed PDFs."
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("source"),
        help="Folder containing source PDF files.",
    )
    parser.add_argument(
        "--processed",
        type=Path,
        default=Path("processed"),
        help="Folder where processed PDF files will be written.",
    )
    parser.add_argument(
        "--project-name",
        required=True,
        help="Project name appended to every output filename.",
    )
    parser.add_argument(
        "--dpi",
        type=int,
        default=300,
        help="Render DPI used when converting PDF pages to images.",
    )
    parser.add_argument(
        "--lang",
        default="en",
        help="PaddleOCR language code, for example en or ar.",
    )
    parser.add_argument(
        "--disable-angle-cls",
        action="store_true",
        help="Disable PaddleOCR angle classification.",
    )
    parser.add_argument(
        "--save-text",
        action="store_true",
        help="Write OCR text to a .txt file beside each generated PDF.",
    )
    parser.add_argument(
        "--ocr-profile",
        choices=["printed", "handwriting", "mixed"],
        default="mixed",
        help="OCR strategy: printed, handwriting, or mixed ensemble.",
    )
    parser.add_argument(
        "--ocr-backend",
        choices=["normal", "ai"],
        default="ai",
        help="Use normal OCR for faster processing or ai OCR for stronger extraction.",
    )
    parser.add_argument(
        "--tesseract-cmd",
        default="",
        help="Optional absolute path to tesseract.exe when using --ocr-backend normal.",
    )
    parser.add_argument(
        "--excel-name",
        default="document_summary.xlsx",
        help="Excel filename written into the processed folder.",
    )
    parser.add_argument(
        "--single-item-per-page",
        action="store_true",
        help="Skip document splitting and treat each PDF page as one document.",
    )
    parser.add_argument(
        "--export-image-mode",
        choices=["original", "enhanced", "both"],
        default="original",
        help="Export original pages, enhanced OCR images, or both.",
    )
    parser.add_argument(
        "--debug-image-dir",
        type=Path,
        default=None,
        help="Optional folder where original and enhanced PNG debug images are saved.",
    )
    parser.add_argument(
        "--handwriting-backend",
        choices=["none", "trocr"],
        default="none",
        help="Optional fallback OCR for handwritten fields.",
    )
    parser.add_argument(
        "--trocr-model",
        default="microsoft/trocr-base-handwritten",
        help="Hugging Face TrOCR model name used when --handwriting-backend trocr.",
    )
    parser.add_argument(
        "--naming-pattern",
        default=DEFAULT_NAMING_PATTERN,
        help=(
            "Output naming pattern using {doc_type}, {date}, {number}, "
            "{company_name}, {amount}, {amount_aed}, {project_name}."
        ),
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    validate_runtime(args.ocr_backend)
    validate_runtime_requirements(
        args.ocr_backend,
        handwriting_backend=args.handwriting_backend,
    )
    configure_model_cache()

    if args.ocr_backend == "normal" and args.tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = args.tesseract_cmd

    source_dir = args.source
    output_dir = args.processed

    if not source_dir.exists():
        raise FileNotFoundError(f"Source folder does not exist: {source_dir}")

    generated_files, records = process_folder(
        source_dir=source_dir,
        output_dir=output_dir,
        project_name=args.project_name,
        dpi=args.dpi,
        language=args.lang,
        use_angle_cls=not args.disable_angle_cls,
        save_text=args.save_text,
        ocr_profile=args.ocr_profile,
        single_item_per_page=args.single_item_per_page,
        ocr_backend=args.ocr_backend,
        export_image_mode=args.export_image_mode,
        debug_image_dir=args.debug_image_dir,
        handwriting_backend=args.handwriting_backend,
        trocr_model=args.trocr_model,
        naming_pattern=args.naming_pattern,
    )
    reconcile_bank_transactions(records)
    write_excel_summary(records, output_dir / args.excel_name)
    print(f"Generated {len(generated_files)} file(s) in {output_dir}")


if __name__ == "__main__":
    main()
